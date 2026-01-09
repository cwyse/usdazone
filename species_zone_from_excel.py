#!/usr/bin/env python3
"""
species_zone_from_excel.py

Read an Excel file with a species column, parse/normalize "Genus species" (supports infra ranks),
dedupe, compute a GBIF-based cold-edge USDA-like zone per unique species, and write an output workbook.

Key properties
--------------
- Resume-safe cache loading from existing --out-xlsx
- Atomic output writes: tmp -> os.replace
- Threaded GBIF fetch only (HTTP-bound)
- Climate dataset opened once for compute phase: with USDAZoneDataset(...) as zds:

Important behavior
------------------
--retry-unknown forces refetch for any cached rows whose zone is Unknown (regardless of reason text),
and also for a set of known failure reasons (normalized).
"""

from __future__ import annotations

import argparse
import os
import random
import re
import sys
import time
import threading
import zlib
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, List, Optional, Sequence, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from usda_zone_access import USDAZoneDataset

from gbif_cold_edge import (
    make_gbif_session,
    ColdEdgeConfig,
    OccPoint,
    fetch_gbif_occurrences,
    estimate_from_occurrences,
    GbifFetchDiagnostics,
    GbifError,
)

# --------------------------------------------------------------------
# Global GBIF throttle (cross-thread) - enforced in this script BEFORE paging starts
# --------------------------------------------------------------------
_GBIF_LOCK = threading.Lock()
_GBIF_NEXT_TS = 0.0

_GBIF_MIN_INTERVAL_S = 0.35  # ~3 req/sec across all threads; tune 0.25..0.7

def _gbif_global_throttle(min_interval_s: float) -> None:
    """Enforce a minimum interval between *any* GBIF requests across threads."""
    global _GBIF_NEXT_TS
    with _GBIF_LOCK:
        now = time.time()
        if now < _GBIF_NEXT_TS:
            time.sleep(_GBIF_NEXT_TS - now)
        _GBIF_NEXT_TS = time.time() + float(min_interval_s)


# --------------------------------------------------------------------
# Species parsing (binomial + infra rank support)
# --------------------------------------------------------------------
GENUS_TOKEN_RE = r"[A-Z][a-zA-Z-]+"
SPECIES_TOKEN_RE = r"[a-z][a-zA-Z-]+"

_BINOMIAL_RE = re.compile(
    rf"^\s*({GENUS_TOKEN_RE})"              # Genus
    rf"(?:\s*\([^)]*\))?"                   # optional "(...)" right after genus
    rf"\s+({SPECIES_TOKEN_RE})\b"           # species epithet
)

_INFRA_RE = re.compile(
    r"\b(ssp|subsp|var|v|f|forma)\.?\s+([a-z][a-zA-Z-]+)\b",
    re.IGNORECASE,
)

def _norm_rank(rank: str) -> str:
    r = rank.lower().strip().strip(".")
    if r in ("ssp", "subsp"):
        return "subsp."
    if r in ("var", "v"):
        return "var."
    if r in ("f", "forma"):
        return "f."
    return r + "."

def parse_species_for_gbif(value: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Returns (parsed_binomial, primary_gbif_query)

    - Parses parenthetical synonyms: "Oreostemma (Aster) alpigenum" -> "Oreostemma alpigenum"
    - Skips genus placeholders: "Oxytropis sp", "Oxytropis sp.", "Oxytropis spp", "Oxytropis spp." -> (None, None)
    - If infraspecific rank is present, uses infra as primary query:
         "Penstemon newberryi ssp newberryi" -> primary query "Penstemon newberryi subsp. newberryi"
      Caller falls back to binomial if infra yields 0 points.
    """
    if value is None:
        return (None, None)

    if isinstance(value, (int, float)):
        # treat numeric cells as invalid species
        return (None, None)

    s = str(value).strip()
    if not s or s == "0":
        return (None, None)

    m = _BINOMIAL_RE.match(s)
    if not m:
        return (None, None)

    genus = m.group(1)
    epithet = m.group(2)

    # Skip placeholders like "Genus sp/sp./spp/spp."
    if epithet.lower() in {"sp", "spp"}:
        return (None, None)

    binomial = f"{genus} {epithet}"

    infra_m = _INFRA_RE.search(s)
    if infra_m:
        rank = _norm_rank(infra_m.group(1))
        infra = infra_m.group(2)
        return (binomial, f"{binomial} {rank} {infra}")

    return (binomial, binomial)


# --------------------------------------------------------------------
# Excel helpers
# --------------------------------------------------------------------
def find_header_column(ws: Worksheet, header_name: str, header_row: int) -> int:
    target = header_name.strip()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        if str(v).strip() == target:
            return col
    raise ValueError(f"Could not find header '{header_name}' in row {header_row}.")

def excel_cell_ref(ws: Worksheet, row: int, col: int) -> str:
    return ws.cell(row=row, column=col).coordinate

def atomic_save_xlsx(wb: Workbook, out_path: str) -> None:
    out_dir = os.path.dirname(os.path.abspath(out_path)) or "."
    base = os.path.basename(out_path)
    tmp = os.path.join(out_dir, f".{base}.tmp.{os.getpid()}.{int(time.time())}")
    wb.save(tmp)
    os.replace(tmp, out_path)

def _norm_text_cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    # normalize NBSP and common weird whitespace
    s = s.replace("\u00a0", " ").strip()
    return s

def _norm_reason(v: Any) -> str:
    s = _norm_text_cell(v)
    return s.upper()

def _is_unknown_zone(v: Any) -> bool:
    s = _norm_text_cell(v)
    return (not s) or (s.lower() == "unknown")

def load_cache_or_empty(
    out_xlsx: str,
    *,
    verbose: bool,
) -> Dict[str, Tuple[Optional[float], Optional[float], str, str, str]]:
    """
    Output workbook format (sheet 1):
      A: parsed_species
      B: latitude
      C: longitude
      D: zone
      E: zone_reason
      F: zone_detail
    """
    cache: Dict[str, Tuple[Optional[float], Optional[float], str, str, str]] = {}
    if not os.path.exists(out_xlsx):
        return cache

    try:
        wb = load_workbook(out_xlsx, data_only=True)
        ws = wb.active

        for r in range(2, ws.max_row + 1):
            sp = ws.cell(row=r, column=1).value
            sp_s = _norm_text_cell(sp)
            if not sp_s:
                continue

            lat = ws.cell(row=r, column=2).value
            lon = ws.cell(row=r, column=3).value
            zone = ws.cell(row=r, column=4).value

            reason = ws.cell(row=r, column=5).value if ws.max_column >= 5 else None
            detail = ws.cell(row=r, column=6).value if ws.max_column >= 6 else None

            latf: Optional[float]
            lonf: Optional[float]
            try:
                latf = float(lat) if lat is not None else None
            except Exception:
                latf = None
            try:
                lonf = float(lon) if lon is not None else None
            except Exception:
                lonf = None

            z = _norm_text_cell(zone) or "Unknown"
            if not z:
                z = "Unknown"

            if reason is None:
                reason_s = "OK" if not _is_unknown_zone(z) else "UNKNOWN_LEGACY"
            else:
                reason_s = _norm_text_cell(reason) or ("OK" if not _is_unknown_zone(z) else "UNKNOWN_LEGACY")

            detail_s = _norm_text_cell(detail)

            cache[sp_s] = (latf, lonf, z, reason_s, detail_s)

        return cache

    except Exception:
        ts = time.strftime("%Y%m%d_%H%M%S")
        bad = f"{out_xlsx}.bad.{ts}"
        os.replace(out_xlsx, bad)
        if verbose:
            print(f"Existing output workbook unreadable/corrupt; renamed to: {bad}", flush=True)
        return {}

def write_output(
    out_xlsx: str,
    parsed_rows: Sequence[Tuple[Optional[str], Any]],
    cache: Dict[str, Tuple[Optional[float], Optional[float], str, str, str]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "species_zones"
    ws.append(["parsed_species", "latitude", "longitude", "zone", "zone_reason", "zone_detail"])

    for parsed, _raw in parsed_rows:
        if not parsed:
            continue
        lat, lon, zone, reason, detail = cache.get(parsed, (None, None, "Unknown", "UNKNOWN_LEGACY", ""))
        ws.append([parsed, lat, lon, zone, reason, detail])

    atomic_save_xlsx(wb, out_xlsx)


# --------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------
def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Compute GBIF cold-edge USDA-like zone per species from an Excel column.")
    p.add_argument("--in-xlsx", required=True)
    p.add_argument("--sheet", default=None)
    p.add_argument("--species-col", required=True)
    p.add_argument("--header-row", type=int, default=1)

    p.add_argument("--dataset", required=True)
    p.add_argument("--out-xlsx", required=True)

    p.add_argument("--workers", type=int, default=3, help="GBIF fetch threads (default: 3).")
    p.add_argument("--save-every", type=int, default=50, help="Save after N newly-computed species (default: 50).")
    p.add_argument("--no-resume", action="store_true")

    p.add_argument("--max-records", type=int, default=5000)
    p.add_argument("--page-size", type=int, default=300)
    p.add_argument("--max-uncertainty-m", type=int, default=10_000)

    p.add_argument("--no-dominance-pruning", action="store_true")
    p.add_argument("--tropics-abs-lat-deg", type=float, default=20.0)

    p.add_argument("--no-thin", action="store_true")
    p.add_argument("--grid-km", type=float, default=25.0)

    p.add_argument("--use-min", action="store_true")
    p.add_argument("--quantile", type=float, default=0.05)
    p.add_argument("--drivers", type=int, default=25)

    p.add_argument("--verbose", action="store_true")
    p.add_argument(
        "--retry-unknown",
        action="store_true",
        help="Recompute species whose cached zone is Unknown (and/or has certain failure reasons).",
    )

    return p


# --------------------------------------------------------------------
# Main
# --------------------------------------------------------------------
def _det_jitter_s(key: str, attempt: int, base: float = 0.35) -> float:
    """
    Deterministic jitter to prevent thread synchronization.
    Produces a value in [0, base).
    """
    h = zlib.crc32(f"{key}|{attempt}".encode("utf-8")) & 0xFFFFFFFF
    return (h / 2**32) * base

def main(argv: List[str]) -> int:
    args = build_arg_parser().parse_args(argv)

    cfg = ColdEdgeConfig(
        max_records=args.max_records,
        page_size=args.page_size,
        max_uncertainty_m=args.max_uncertainty_m,
        dominance_pruning=not args.no_dominance_pruning,
        tropics_abs_lat_deg=args.tropics_abs_lat_deg,
        do_thin=not args.no_thin,
        grid_km=args.grid_km,
        use_min=args.use_min,
        quantile=args.quantile,
        drivers=args.drivers,
    )

    wb_in = load_workbook(args.in_xlsx, data_only=True)
    ws_in = wb_in[args.sheet] if args.sheet else wb_in.active
    sp_col = find_header_column(ws_in, args.species_col, args.header_row)

    parsed_rows: List[Tuple[Optional[str], Any]] = []
    unique: List[str] = []
    seen: set[str] = set()
    species_to_query: Dict[str, str] = {}

    total_rows = max(0, ws_in.max_row - args.header_row)
    for i, r in enumerate(range(args.header_row + 1, ws_in.max_row + 1), start=1):
        raw = ws_in.cell(row=r, column=sp_col).value
        parsed, query = parse_species_for_gbif(raw)
        parsed_rows.append((parsed, raw))

        if args.verbose:
            print(
                f"[{i}/{total_rows}] {excel_cell_ref(ws_in, r, sp_col)} raw={raw!r} parsed={parsed!r}",
                flush=True,
            )

        if parsed and parsed not in seen:
            seen.add(parsed)
            unique.append(parsed)
            if query:
                species_to_query[parsed] = query

    cache: Dict[str, Tuple[Optional[float], Optional[float], str, str, str]] = {}
    if not args.no_resume:
        cache = load_cache_or_empty(args.out_xlsx, verbose=args.verbose)

    # Always create/update output early
    write_output(args.out_xlsx, parsed_rows, cache)
    if args.verbose:
        print(f"Wrote initial output workbook: {args.out_xlsx}", flush=True)

    # Determine which species need computation/refetch
    retry_reasons = {
        "UNKNOWN_LEGACY",
        "GBIF_ERROR",
        "NO_GBIF_POINTS",
        "NO_CLIMATE_SAMPLES",
        "ESTIMATE_ERROR",
        "GBIF_RATE_LIMIT",
        "GBIF_TIMEOUT",
        "GBIF_CONNECTION_ERROR",
        "GBIF_SERVER_ERROR",
        "GBIF_HTTP_ERROR",
    }

    need: List[str] = []
    retry_breakdown: Dict[str, int] = {}

    for sp in unique:
        if sp not in cache:
            need.append(sp)
            retry_breakdown["NOT_IN_CACHE"] = retry_breakdown.get("NOT_IN_CACHE", 0) + 1
            continue

        if args.retry_unknown:
            _lat, _lon, zone, reason, _detail = cache[sp]
            zone_unknown = _is_unknown_zone(zone)
            reason_norm = _norm_reason(reason)

            # Force refetch if zone is Unknown, regardless of reason text.
            if zone_unknown:
                need.append(sp)
                retry_breakdown["ZONE_UNKNOWN"] = retry_breakdown.get("ZONE_UNKNOWN", 0) + 1
                continue

            # Also refetch for known failure reasons (covers non-Unknown zone cases or legacy inconsistency)
            if reason_norm in retry_reasons:
                need.append(sp)
                retry_breakdown[reason_norm] = retry_breakdown.get(reason_norm, 0) + 1

    if args.verbose:
        print(f"Unique parsed species: {len(unique)} (need compute: {len(need)}, cached: {len(cache)})", flush=True)
        if args.retry_unknown:
            items = sorted(retry_breakdown.items(), key=lambda t: (-t[1], t[0]))
            print("Retry breakdown:", ", ".join(f"{k}={v}" for k, v in items) or "(none)", flush=True)

    # Verify intended import path
    if args.verbose:
        import gbif_cold_edge as gce
        print(f"Using gbif_cold_edge.py from: {gce.__file__}", flush=True)

    # Thread-local sessions
    _thread_local = threading.local()

    def _get_thread_session() -> requests.Session:
        sess = getattr(_thread_local, "sess", None)
        if sess is None:
            sess = make_gbif_session(pool_connections=2, pool_maxsize=2)
            _thread_local.sess = sess
        return sess

    def fetch_one(sp: str, query: str) -> Tuple[str, List[OccPoint], Optional[str], str, str]:
        """
        Returns:
          (parsed_species_key, occurrences, err_string_or_None, used_query, diag_detail)
        """
        diag = GbifFetchDiagnostics()
        MIN_INTERVAL_S = _GBIF_MIN_INTERVAL_S
        MAX_TRIES = 12
        BASE_BACKOFF = 1.7

        last_err: Optional[str] = None
        last_diag_detail = ""

        for attempt in range(1, MAX_TRIES + 1):
            try:
                sess = _get_thread_session()

                # Global throttle BEFORE entering GBIF paging loop
                _gbif_global_throttle(MIN_INTERVAL_S)

                pts = fetch_gbif_occurrences(query, cfg, session=sess, diag=diag)

                diag_detail = (
                    f"mode={diag.used_query_mode} usageKey={diag.usage_key} "
                    f"pages={diag.pages} total={diag.gbif_total_reported} "
                    f"seen={diag.n_raw_records_seen} kept={diag.n_kept} "
                    f"miss_coord={diag.n_missing_coords} "
                    f"unc_filt={diag.n_uncertainty_filtered} "
                    f"basis_filt={diag.n_basis_filtered} "
                    f"est_filt={diag.n_establishment_filtered}"
                )

                return sp, pts, None, query, diag_detail

            except GbifError as e:
                reason = getattr(e, "reason", "GBIF_ERROR")
                detail = getattr(e, "detail", str(e))
                last_err = f"{reason}: {detail}"

                last_diag_detail = (
                    f"mode={diag.used_query_mode} usageKey={diag.usage_key} "
                    f"pages={diag.pages} total={diag.gbif_total_reported} "
                    f"seen={diag.n_raw_records_seen} kept={diag.n_kept} "
                    f"fail={diag.fail_reason}:{diag.fail_detail}"
                )

                # Respect Retry-After when present; otherwise exponential with deterministic jitter
                sleep_s: float
                if reason == "GBIF_RATE_LIMIT":
                    m = re.search(r"Retry-After=(\d+)", str(detail))
                    if m:
                        sleep_s = float(m.group(1)) + _det_jitter_s(sp, attempt, base=0.5)
                    else:
                        sleep_s = min(60.0, (BASE_BACKOFF ** (attempt - 1)) + _det_jitter_s(sp, attempt, base=0.6))
                else:
                    sleep_s = min(30.0, (BASE_BACKOFF ** (attempt - 1)) + _det_jitter_s(sp, attempt, base=0.4))

                time.sleep(max(1.0, sleep_s))
                continue

            except Exception as e:
                last_err = f"{type(e).__name__}: {e}"
                last_diag_detail = (
                    f"mode={diag.used_query_mode} usageKey={diag.usage_key} "
                    f"pages={diag.pages} total={diag.gbif_total_reported} "
                    f"seen={diag.n_raw_records_seen} kept={diag.n_kept} "
                    f"fail={type(e).__name__}"
                )
                sleep_s = min(30.0, (BASE_BACKOFF ** (attempt - 1)) + _det_jitter_s(sp, attempt, base=0.4))
                time.sleep(max(1.0, sleep_s))
                continue

        return sp, [], last_err or "GBIF_ERROR: exhausted retries", query, last_diag_detail

    fetched: Dict[str, List[OccPoint]] = {}
    fetch_errors: Dict[str, str] = {}
    fetch_diag: Dict[str, str] = {}

    workers = max(1, min(int(args.workers), 4))

    if need:
        start_t = time.time()
        last_beat = start_t

        if args.verbose:
            print(f"Starting GBIF fetch: need={len(need)} workers={workers} throttle={_GBIF_MIN_INTERVAL_S:.2f}s", flush=True)

        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs: Dict[Any, Tuple[str, str]] = {}

            for sp in need:
                q = species_to_query.get(sp, sp)
                futs[ex.submit(fetch_one, sp, q)] = (sp, q)

            for fut in as_completed(futs):
                sp_expected, q_expected = futs[fut]

                try:
                    sp2, pts, err, used_query, diag_detail = fut.result()
                except Exception as e:
                    fetched[sp_expected] = []
                    fetch_errors[sp_expected] = f"future_exception: {type(e).__name__}: {e}"
                    continue

                sp_key = sp_expected

                # If infra query returns 0 points, fall back to binomial automatically
                if not err and not pts and used_query != sp_key:
                    if args.verbose:
                        print(f"  .. GBIF empty infra='{used_query}', retry binomial='{sp_key}'", flush=True)
                    _spb, pts2, err2, _used2, diag2 = fetch_one(sp_key, sp_key)
                    pts, err, diag_detail = pts2, err2, diag2

                fetch_diag[sp_key] = (diag_detail or "")[:250]
                fetched[sp_key] = pts

                if err:
                    fetch_errors[sp_key] = err

                now = time.time()
                if args.verbose and (now - last_beat) >= 15:
                    remaining = sum(1 for f2 in futs if not f2.done())
                    err_ct = len(fetch_errors)
                    print(f"  .. GBIF fetch heartbeat: remaining={remaining} errors={err_ct}", flush=True)
                    last_beat = now

        if args.verbose:
            # quick histogram of top error reasons
            hist: Dict[str, int] = {}
            for _sp, msg in fetch_errors.items():
                reason = msg.split(":", 1)[0].strip()
                hist[reason] = hist.get(reason, 0) + 1
            top = sorted(hist.items(), key=lambda t: (-t[1], t[0]))[:10]
            if top:
                print("GBIF error histogram (top): " + ", ".join(f"{k}={v}" for k, v in top), flush=True)

    # Compute stage (single-thread, dataset open once)
    last_save_t = time.time()
    last_prog_t = time.time()

    computed = 0
    with USDAZoneDataset(args.dataset) as zds:
        for sp in need:
            if sp in fetch_errors:
                cache[sp] = (None, None, "Unknown", "GBIF_ERROR",
                             f"{fetch_errors[sp][:160]} | {fetch_diag.get(sp, '')}"[:250])
            else:
                occs = fetched.get(sp, [])
                if not occs:
                    cache[sp] = (None, None, "Unknown", "NO_GBIF_POINTS", fetch_diag.get(sp, "")[:250])
                else:
                    try:
                        res = estimate_from_occurrences(
                            species=sp,
                            dataset_path=args.dataset,
                            cfg=cfg,
                            occurrences=occs,
                            zds=zds,
                        )
                        edge_cp, edge_occ = res.cold_edge
                        zone_val = getattr(edge_cp, "zone", None) or getattr(edge_cp, "zone_label", None) or "Unknown"
                        cache[sp] = (float(edge_occ.lat), float(edge_occ.lon), str(zone_val), "OK", fetch_diag.get(sp, "")[:250])
                    except Exception as e:
                        msg = str(e)
                        if isinstance(e, RuntimeError) and msg.startswith("No climate samples."):
                            reason = "NO_CLIMATE_SAMPLES"
                        else:
                            reason = "ESTIMATE_ERROR"
                        cache[sp] = (None, None, "Unknown", reason, msg[:250])

            computed += 1

            now = time.time()
            if args.verbose and (now - last_prog_t) >= 30:
                print(f"  .. compute heartbeat: {computed}/{len(need)} computed", flush=True)
                last_prog_t = now

            if args.save_every > 0 and computed % int(args.save_every) == 0:
                write_output(args.out_xlsx, parsed_rows, cache)
                last_save_t = time.time()
                if args.verbose:
                    print(f"  .. saved checkpoint ({computed} computed) -> {args.out_xlsx}", flush=True)

            if (time.time() - last_save_t) >= 300:
                write_output(args.out_xlsx, parsed_rows, cache)
                last_save_t = time.time()
                if args.verbose:
                    print(f"  .. saved timed checkpoint -> {args.out_xlsx}", flush=True)

    write_output(args.out_xlsx, parsed_rows, cache)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
