#!/usr/bin/env python3
"""
species_inat_links_from_excel_clickable.py

Read an Excel file with a species column, parse/normalize binomial names, and write an
output workbook that MATCHES the INPUT ROW COUNT (one output row per input row, in the
same order).

Primary output sheet:
  - Sheet: inat_links_rows
  - One row per input row (excluding header row), including:
      excel_row (Excel row #), raw_value, parsed_species, inat_map_url

Optional convenience sheet:
  - Sheet: inat_links_unique
  - One row per unique parsed_species (deduped), including:
      parsed_species, inat_map_url, example_raw_input, first_excel_row

Hyperlinks:
  - The inat_map_url cells are written as real Excel hyperlinks (clickable).
  - The displayed cell text remains the URL string (or "Unknown" / blank).

Behavior for duplicates:
  - Default: duplicates are repeated with the same URL.
  - If you prefer duplicates to be blanked after first occurrence, use:
      --blank-duplicates

iNaturalist URL behavior:
  - Uses taxon_name=... (species name), not taxon_id.
  - Uses verifiable=any and a world-scale bbox to force map fully zoomed out:
      nelat=85 nelng=180 swlat=-85 swlng=-180
  - Does NOT call iNaturalist APIs; it only generates deterministic URLs.

Example:
  python ./species_inat_links_from_excel_clickable.py \
    --in-xlsx your_species.xlsx \
    --species-col "Species" \
    --out-xlsx inat_links.xlsx \
    --verbose
"""

from __future__ import annotations

import argparse
import os
import re
import time
import urllib.parse
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


# ----------------------------
# Species parsing (compatible with earlier pipeline)
# ----------------------------

GENUS_TOKEN_RE = r"[A-Z][a-zA-Z-]+"
SPECIES_TOKEN_RE = r"[a-z][a-zA-Z-]+"

_BINOMIAL_RE = re.compile(
    rf"^\s*({GENUS_TOKEN_RE})"              # Genus
    rf"(?:\s*\([^)]*\))?"                 # optional "(...)" right after genus
    rf"\s+({SPECIES_TOKEN_RE})\b"          # species epithet
)

_INFRA_RE = re.compile(
    r"\b(ssp|subsp|var|v|f|forma)\.?\s+([a-z][a-zA-Z-]+)\b",
    re.IGNORECASE,
)


def _norm_rank(rank: str) -> str:
    r = rank.lower().strip().strip(".")
    if r in ("ssp", "subsp"):
        return "subsp."
    if r in ("var", "v"):
        return "var."
    if r in ("f", "forma"):
        return "f."
    return r + "."


def parse_species_for_gbif(value: Any) -> Tuple[Optional[str], Optional[str]]:
    """Return (parsed_binomial, primary_query_or_None)."""
    if value is None:
        return (None, None)

    if isinstance(value, (int, float)):
        try:
            if float(value) == 0.0:
                return (None, None)
        except Exception:
            return (None, None)
        return (None, None)

    s = str(value).strip()
    if not s or s == "0":
        return (None, None)

    m = _BINOMIAL_RE.match(s)
    if not m:
        return (None, None)

    genus = m.group(1)
    epithet = m.group(2)

    if epithet.lower() in {"sp", "spp"}:
        return (None, None)

    binomial = f"{genus} {epithet}"

    infra_m = _INFRA_RE.search(s)
    if infra_m:
        rank = _norm_rank(infra_m.group(1))
        infra = infra_m.group(2)
        return (binomial, f"{binomial} {rank} {infra}")

    return (binomial, binomial)


# ----------------------------
# Excel helpers
# ----------------------------

def find_header_column(ws: Worksheet, header_name: str, header_row: int) -> int:
    target = header_name.strip()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        if str(v).strip() == target:
            return col
    raise ValueError(f"Could not find header '{header_name}' in row {header_row}.")


def atomic_save_xlsx(wb: Workbook, out_path: str) -> None:
    out_dir = os.path.dirname(os.path.abspath(out_path)) or "."
    base = os.path.basename(out_path)
    tmp = os.path.join(out_dir, f".{base}.tmp.{os.getpid()}.{int(time.time())}")
    wb.save(tmp)
    os.replace(tmp, out_path)


# ----------------------------
# iNaturalist URL builder
# ----------------------------

def inat_map_global_by_name(species_name: str) -> str:
    q = urllib.parse.quote_plus(species_name)
    return (
        "https://www.inaturalist.org/observations"
        f"?taxon_name={q}"
        "&verifiable=any"
        "&subview=map"
        "&nelat=85&nelng=180"
        "&swlat=-85&swlng=-180"
    )


def _set_hyperlink(cell, url: str) -> None:
    """Make a cell clickable if url looks like an http(s) URL."""
    if not url:
        return
    if not (url.startswith("http://") or url.startswith("https://")):
        return
    cell.hyperlink = url
    cell.font = Font(color="0000FF", underline="single")


@dataclass
class RowOut:
    excel_row: int
    raw_value: Any
    parsed_species: Optional[str]
    inat_url: str


# ----------------------------
# CLI
# ----------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Create iNaturalist map links for species listed in an Excel column.")
    p.add_argument("--in-xlsx", required=True)
    p.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    p.add_argument("--species-col", required=True, help="Header text of the species column")
    p.add_argument("--header-row", type=int, default=1)
    p.add_argument("--out-xlsx", required=True)
    p.add_argument(
        "--blank-duplicates",
        action="store_true",
        help="If set, only the first occurrence of a parsed species gets a URL; later duplicates are blank.",
    )
    p.add_argument("--verbose", action="store_true")
    return p


def main(argv: List[str]) -> int:
    args = build_arg_parser().parse_args(argv)

    wb_in = load_workbook(args.in_xlsx, data_only=True)
    ws_in = wb_in[args.sheet] if args.sheet else wb_in.active
    sp_col = find_header_column(ws_in, args.species_col, args.header_row)

    rows_out: List[RowOut] = []

    first_seen_row: Dict[str, int] = {}
    example_raw_for_species: Dict[str, str] = {}
    seen_species: set[str] = set()

    total_rows = max(0, ws_in.max_row - args.header_row)

    for i, excel_r in enumerate(range(args.header_row + 1, ws_in.max_row + 1), start=1):
        raw = ws_in.cell(row=excel_r, column=sp_col).value
        parsed, _q = parse_species_for_gbif(raw)

        if parsed:
            if parsed not in first_seen_row:
                first_seen_row[parsed] = excel_r
                example_raw_for_species[parsed] = "" if raw is None else str(raw)

            if args.blank_duplicates and parsed in seen_species:
                url = ""
            else:
                url = inat_map_global_by_name(parsed)

            seen_species.add(parsed)
        else:
            url = "Unknown"

        rows_out.append(RowOut(excel_row=excel_r, raw_value=raw, parsed_species=parsed, inat_url=url))

        if args.verbose:
            shown = "(blank)" if url == "" else url
            print(f"[{i}/{total_rows}] row={excel_r} raw={raw!r} parsed={parsed!r} url={shown}", flush=True)

    wb_out = Workbook()

    # Sheet 1: row-for-row output
    ws_rows = wb_out.active
    ws_rows.title = "inat_links_rows"
    ws_rows.append(["excel_row", "raw_value", "parsed_species", "inat_map_url"])

    for ro in rows_out:
        ws_rows.append([
            ro.excel_row,
            "" if ro.raw_value is None else str(ro.raw_value),
            ro.parsed_species or "",
            ro.inat_url,
        ])
        url_cell = ws_rows.cell(row=ws_rows.max_row, column=4)
        _set_hyperlink(url_cell, ro.inat_url)

    # Sheet 2: unique species convenience view
    ws_unique = wb_out.create_sheet("inat_links_unique")
    ws_unique.append(["parsed_species", "inat_map_url", "example_raw_input", "first_excel_row"])

    unique_sorted = sorted(first_seen_row.items(), key=lambda kv: kv[1])
    for sp, first_row in unique_sorted:
        url = inat_map_global_by_name(sp)
        ws_unique.append([sp, url, example_raw_for_species.get(sp, ""), first_row])
        url_cell = ws_unique.cell(row=ws_unique.max_row, column=2)
        _set_hyperlink(url_cell, url)

    # Sheet 3: unparsed inputs (optional)
    unparsed = [ro for ro in rows_out if ro.parsed_species is None and ro.raw_value not in (None, "", 0, "0")]
    if unparsed:
        ws_un = wb_out.create_sheet("unparsed_inputs")
        ws_un.append(["excel_row", "raw_value", "reason"])
        for ro in unparsed:
            ws_un.append([ro.excel_row, "" if ro.raw_value is None else str(ro.raw_value), "could not parse binomial"])

    atomic_save_xlsx(wb_out, args.out_xlsx)

    if args.verbose:
        print(f"Wrote: {args.out_xlsx}")
        print(f"Input rows processed: {len(rows_out)}")
        print(f"Unique parsed species: {len(first_seen_row)}")
        if unparsed:
            print(f"Unparsed inputs: {len(unparsed)} (see sheet 'unparsed_inputs')")
        if args.blank_duplicates:
            print("Duplicate behavior: URLs blanked after first occurrence.")
        else:
            print("Duplicate behavior: URLs repeated for each occurrence (row-for-row mapping).")

    return 0


if __name__ == "__main__":
    import sys
    raise SystemExit(main(sys.argv[1:]))
